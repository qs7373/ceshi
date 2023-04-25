import sqlite3
import requests
from tkinter import ttk
from tkinter.filedialog import askopenfilename
import os
import pickle
import subprocess
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader,TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from PyPDF2 import PdfFileWriter, PdfFileReader
from reportlab.lib.units import mm, inch
from PyQt5.QtWidgets import QApplication, QWidget, QLabel, QVBoxLayout
from PyQt5.QtGui import QFontMetrics
import pandas as pd
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
import bluetooth
import cups
from PyQt5.QtCore import Qt, QPoint
from PyQt5.QtGui import QPen, QPainter, QImage, QPixmap
from PyQt5.QtWidgets import QWidget, QLabel, QVBoxLayout, QHBoxLayout, QComboBox, QFileDialog, QPushButton, QSizePolicy
from tkinter import *
from reportlab.lib.pagesizes import letter, legal
from reportlab.lib.units import inch
from PIL import Image
from reportlab.lib.pagesizes import letter, legal, tabloid, A5, A4, A3, B5, B4, B3


# 创建数据库连接
conn = sqlite3.connect("cards.db")
c = conn.cursor()

# 创建cards表
c.execute("""CREATE TABLE IF NOT EXISTS cards (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                content TEXT NOT NULL,
                image_path TEXT,
                paper_size TEXT NOT NULL,
                font TEXT NOT NULL,
                size INTEGER NOT NULL,
                leading REAL NOT NULL,
                tracking REAL NOT NULL,
                crop_margin REAL NOT NULL,
                vertical BOOLEAN NOT NULL
            )""")
conn.commit()



# 向指定URL发送POST请求
def send_post_request(url, data):
    response = requests.post(url, data=data)
    return response.text

# 定义变量和全局变量
font_sizes = ["8", "9", "10", "11", "12", "14", "16", "18", "20", "22", "24", "28", "32", "36", "40", "44", "48"]
fonts_dir = "C:\\Windows\\Fonts\\"
system_fonts = []
selected_image = None
settings = {"size": "A4", "custom_size": "", "font": "SimSun", "size_num": 12, "leading": 12, "tracking": 0}
card_settings = {
    "paper_size": "A4",
    "custom_size": "210x297",
    "font": "SimSun",
    "size": 24,
    "leading": 36,
    "crop_margin": 10,
    "auto_crop": True,
    "pdf_filename": "cards.pdf",
    "image_filename": None,
    "text1": "佛光注照\n长生禄位",
    "text2": "佛力超荐\n往生莲位\n阳上:\n拜荐"
}
card_settings = {
    "paper_size": tk.StringVar(),
    "font": tk.StringVar(),
    "size": tk.IntVar(),
    "leading": tk.DoubleVar(),
    "tracking": tk.DoubleVar(),
    "crop_margin": tk.DoubleVar(),
    "vertical": tk.BooleanVar()
}
card_settings["paper_size"].set("A4")
card_settings["font"].set("SimSun")
card_settings["size"].set(20)
card_settings["leading"].set(1.5)
card_settings["tracking"].set(0)
card_settings["crop_margin"].set(3)
card_settings["vertical"].set(False)

# 定义常量
PAPER_SIZES = {
    "A0": (841, 1189),
    "A1": (594, 841),
    "A2": (420, 594),
    "A3": (297, 420),
    "A4": (210, 297),
    "A5": (148, 210),
    "A6": (105, 148),
    "A7": (74, 105),
    "A8": (52, 74),
    "A9": (37, 52),
    "A10": (26, 37)
}

# 创建GUI窗口
root = Tk()
root.title("牌位打印")
root.geometry("800x600")

class App:
    def __init__(self, root):
        self.root = root
        self.root.title("牌位打印程序")
        self.root.geometry("800x600")
        self.root.resizable(False, False)
        self.master = self.root
        self.create_widgets()
        self.master = self.root

        # 从配置文件中读取设置
        try:
            with open("settings.pkl", "rb") as f:
                settings = pickle.load(f)
            self.size_var.set(settings["size"])
            self.custom_size_entry.delete(0, END)
            self.custom_size_entry.insert(0, settings["custom_size"])
            self.font_entry.delete(0, END)
            self.font_entry.insert(0, settings["font"])
            self.size_entry.delete(0, END)
            self.size_entry.insert(0, settings["size"])
            self.leading_entry.delete(0, END)
            self.leading_entry.insert(0, settings["leading"])
            self.tracking_entry.delete(0, END)
            self.tracking_entry.insert(0, settings["tracking"])
        except FileNotFoundError:
            pass

# 添加图片
logo_file = "logo.png"
if os.path.exists(logo_file):
    logo_image = PhotoImage(file=logo_file)
    logo_label = Label(image=logo_image)
    logo_label.pack()

# 创建一个Notebook，用于切换不同的功能页
    self.notebook = ttk.Notebook(self.root)
    self.notebook.pack(expand=True, fill=BOTH)

    # 创建一个Frame，用于放置设置页的部件
    self.settings_frame = Frame(self.notebook)
    self.notebook.add(self.settings_frame, text="设置")

    # 创建一个Frame，用于放置打印页的部件
    self.notebook.add(self.print_frame, text="打印")

    # 创建设置页的部件
    self.create_settings_widgets()

    # 创建打印页的部件
    self.create_print_widgets()


def create_settings_widgets(self):
    # 创建设置页的部件
    pass  # 这里原本的pass语句应该删除或者移到函数内部


def create_print_widgets(self):
    # 创建打印页的部件
    file_label = Label(self.print_frame, text="选择PDF文件：")
    file_label.pack()

    # 添加文件路径输入框和选择文件按钮
    self.file_path_label = Label(self.print_frame, text="选择要打印的文件")
    self.file_path_label.pack(pady=10)
    self.file_path_entry = Entry(self.print_frame, width=50)
    self.file_path_entry.pack(pady=10)
    self.file_path_button = Button(self.print_frame, text="选择文件", command=self.choose_file)
    self.file_path_button.pack(pady=10)

    self.print_options_label = Label(self.print_frame, text="打印选项：")
    self.print_options_label.pack(pady=10)

    self.paper_size_label = Label(self.print_frame, text="纸张尺寸：")
    self.paper_size_label.pack()
    self.size_var = StringVar()
    self.size_var.set("A4")
    self.size_combobox = ttk.Combobox(self.print_frame, textvariable=self.size_var, values=["A4", "A3", "A5", "B4", "B3", "B5", "Letter", "Legal", "Tabloid", "Custom"])
    self.size_combobox.pack(pady=5)
    self.size_combobox.bind("<<ComboboxSelected>>", self.on_size_select)

    self.custom_size_label = Label(self.print_frame, text="自定义尺寸（毫米）：")
    self.custom_size_label.pack()
    self.custom_size_entry = Entry(self.print_frame, width=10)
    self.custom_size_entry.pack(pady=5)
    self.custom_size_entry.insert(0, "210x297")

    self.font_label = Label(self.print_frame, text="字体：")
    self.font_label.pack()
    self.font_entry = Entry(self.print_frame, width=20)
    self.font_entry.pack(pady=5)
    self.font_entry.insert(0, "SimSun")

    self.size_label = Label(self.print_frame, text="字号：")
    self.size_label.pack()
    self.size_entry = Entry(self.print_frame, width=5)
    self.size_entry.pack(pady=5)
    self.size_entry.insert(0, "12")

    self.leading_label = Label(self.print_frame, text="行距：")
    self.leading_label.pack()
    self.leading_entry = Entry(self.print_frame, width=5)
    self.leading_entry.pack(pady=5)
    self.leading_entry.insert(0, "20")

    self.tracking_label = Label(self.print_frame, text="字间距：")
    self.tracking_label.pack()
    self.tracking_entry = Entry(self.print_frame, width=5)
    self.tracking_entry.pack(pady=5)
    self.tracking_entry.insert(0, "0")

    self.preview_button = Button(self.print_frame, text="预览", command=self.preview_card)
    self.preview_button.pack(pady=10)

    self.print_button = Button(self.print_frame, text="打印", command=self.print_card)
    self.print_button.pack(pady=10)

def choose_file(self):
    # 选择文件并更新文件路径
    filename = askopenfilename(filetypes=[("PDF files", "*.pdf")])
    self.file_path_entry.delete(0, END)
    self.file_path_entry.insert(0, filename)

def preview_card(self):

     # 预览牌位
    pass


def print_card(self):
    # 打印牌位
    pass

# 在主界面添加文本框
content_label = Label(root, text="输入要打印的内容")
content_label.pack()

# 消灾延生牌位
content1_label = Label(root, text="消灾延生牌位")
content1_label.pack()

content1_text = Text(root, height=3, width=50)
content1_text.insert(END, "佛光注照 () 长生禄位")
content1_text.pack()

# 往生超度牌位
content2_label = Label(root, text="往生超度牌位")
content2_label.pack()

content2_text = Text(root, height=4, width=50)
content2_text.insert(END, "佛力超荐 () 往生莲位\n阳上: () 拜荐")
content2_text.pack()

# 添加设置和打印页面切换按钮
settings_button = Button(root, text="设置", command=settings_page)
settings_button.pack(side=LEFT)

print_button = Button(root, text="打印", command=print_card)
print_button.pack(side=RIGHT)

# 添加自定义牌位模板
class TemplateEditor(QWidget):
    def __init__(self):
        super().__init__()

        self.image = QImage(800, 600, QImage.Format_ARGB32)
        self.image.fill(Qt.white)

        self.label = QLabel()
        self.label.setPixmap(QPixmap.fromImage(self.image))
        self.label.setAlignment(Qt.AlignCenter)

        self.color = Qt.black
        self.pen_width = 2

        self.draw_mode = "line"

        self.init_ui()

    def init_ui(self):
        color_combo = QComboBox()
        color_combo.addItem("Black", Qt.black)
        color_combo.addItem("Red", Qt.red)
        color_combo.addItem("Blue", Qt.blue)
        color_combo.addItem("Green", Qt.green)
        color_combo.addItem("Yellow", Qt.yellow)
        color_combo.currentIndexChanged.connect(self.color_changed)

        pen_width_combo = QComboBox()
        pen_width_combo.addItem("2", 2)
        pen_width_combo.addItem("4", 4)
        pen_width_combo.addItem("6", 6)
        pen_width_combo.addItem("8", 8)
        pen_width_combo.currentIndexChanged.connect(self.pen_width_changed)

        line_button = QPushButton("Line")
        line_button.clicked.connect(self.line_mode)
        rect_button = QPushButton("Rectangle")
        rect_button.clicked.connect(self.rect_mode)
        ellipse_button = QPushButton("Ellipse")
        ellipse_button.clicked.connect(self.ellipse_mode)

        clear_button = QPushButton("Clear")
        clear_button.clicked.connect(self.clear_image)

        save_button = QPushButton("Save")
        save_button.clicked.connect(self.save_image)

        button_layout = QHBoxLayout()
        button_layout.addWidget(line_button)
        button_layout.addWidget(rect_button)
        button_layout.addWidget(ellipse_button)
        button_layout.addWidget(clear_button)
        button_layout.addWidget(save_button)

        color_layout = QHBoxLayout()
        color_layout.addWidget(QLabel("Color:"))
        color_layout.addWidget(color_combo)
        color_layout.addWidget(QLabel("Pen width:"))
        color_layout.addWidget(pen_width_combo)

        layout = QVBoxLayout()
        layout.addWidget(self.label)
        layout.addLayout(color_layout)
        layout.addLayout(button_layout)

        self.setLayout(layout)

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.drawImage(0, 0, self.image)

    def mousePressEvent(self, event):
        self.last_point = event.pos()

    def mouseMoveEvent(self, event):
        painter = QPainter(self.image)
        painter.setPen(QPen(self.color, self.pen_width))
        if self.draw_mode == "line":
            painter.drawLine(self.last_point, event.pos())
        elif self.draw_mode == "rect":
            painter.drawRect(self.last_point.x(), self.last_point.y(), event.pos().x() - self.last_point.x(),
                              event.pos().y() - self.last_point.y())
        elif self.draw_mode == "ellipse":
            painter.drawEllipse(QPoint((self.last_point.x() + event.pos().x()) / 2,
                                        (self.last_point.y() + event.pos().y()) / 2),
                                  abs(event.pos().x() - self.last_point.x()) / 2,
                                  abs(event.pos().y() - self.last_point.y()) / 2)
        self.last_point = event.pos()
        self.update()

    def color_changed(self, index):
        self.color = self.sender().currentData()

    def pen_width_changed(self, index):
        self.pen_width = self.sender().currentData()

# 导入Excel文件
from openpyxl import load_workbook

def import_document():
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx"), ("PDF Files", "*.pdf"), ("Text Files", "*.txt")])
    if file_path:
        if file_path.endswith(".xlsx"):
            # 加载Excel文件
            wb = load_workbook(filename=file_path)
            sheet_name = wb.sheetnames[0]
            sheet = wb[sheet_name]
            rows = sheet.max_row
            cols = sheet.max_column
            # 显示导入的工作表并允许用户选择打印范围
            top = Toplevel(root)
            top.title("选择打印范围")
            Label(top, text=f"{file_path} 选定的工作表: {sheet_name}，共 {rows} 行 {cols} 列").pack()
            range_frame = Frame(top)
            range_frame.pack(padx=10, pady=10)
            range_label = Label(range_frame, text="选择打印范围，如：A1:B10")
            range_label.grid(row=0, column=0, padx=5, pady=5)
            range_entry = Entry(range_frame)
            range_entry.grid(row=0, column=1, padx=5, pady=5)
            ok_button = Button(top, text="确定", command=lambda: print_document(file_path, sheet_name, range_entry.get()))
            ok_button.pack()
        else:
            # 显示导入的文档并允许用户选择打印范围
            top = Toplevel(root)
            top.title("选择打印范围")
            Label(top, text=f"{file_path}").pack()
            range_frame = Frame(top)
            range_frame.pack(padx=10, pady=10)
            range_label = Label(range_frame, text="选择打印范围，如：1-5")
            range_label.grid(row=0, column=0, padx=5, pady=5)
            range_entry = Entry(range_frame)
            range_entry.grid(row=0, column=1, padx=5, pady=5)
            ok_button = Button(top, text="确定", command=lambda: print_document(file_path, "", range_entry.get()))
            ok_button.pack()

# 打印导入的文档

def print_document(file_path, sheet_name, print_range):
    # 确定文件类型
    file_type = os.path.splitext(file_path)[1]
    if file_type == ".xlsx":
        # 加载Excel文件和选择的工作表
        wb = load_workbook(filename=file_path)
        sheet = wb[sheet_name]
        # 确定要打印的范围
        if print_range:
            print_range = print_range.upper()
            if ":" in print_range:
                start_cell, end_cell = print_range.split(":")
                start_col = ord(start_cell[0]) - 65
                start_row = int(start_cell[1:])
                end_col = ord(end_cell[0]) - 65
                end_row = int(end_cell[1:])
                print_area = sheet.iter_rows(min_row=start_row, min_col=start_col, max_row=end_row, max_col=end_col)
            else:
                row_list = print_range.split(",")
                row_list = [int(x) for x in row_list]
                print_area = sheet.iter_rows(min_row=min(row_list), max_row=max(row_list))
        else:
            # 如果没有选择范围，则打印整个工作表
            print_area = sheet.iter_rows()
        # 创建一个临时PDF文件来存储打印内容
        pdf_file = os.path.join(os.path.dirname(file_path), "temp.pdf")
        # 创建一个画布对象将内容打印到PDF文件
        c = canvas.Canvas(pdf_file, pagesize=(8.5 * inch, 11 * inch))
        for row in print_area:
            for cell in row:
                value = cell.value
                if value:
                    c.drawString(cell.column_letter + str(cell.row), str(value))
        c.save()
        # 使用subprocess将PDF文件打印到默认打印机
        subprocess.Popen(["lpr", pdf_file], shell=True)
        # 删除临时PDF文件
        os.remove(pdf_file)
    elif file_type == ".pdf":
        # 加载PDF文件
        input_pdf = PdfFileReader(open(file_path, "rb"))
        # 确定要打印的范围
        if print_range:
            start_page, end_page = print_range.split("-")
            start_page = int(start_page) - 1
            end_page = int(end_page)
            print_area = input_pdf.pages[start_page:end_page]
        else:
            # 如果没有选择范围，则打印整个文档
            print_area = input_pdf.pages
        # 创建一个新的PDF writer并添加所选页面
        output_pdf = PdfFileWriter()
        for page in print_area:
            output_pdf.addPage(page)
        # 创建一个临时PDF文件来存储打印内容
        pdf_file = os.path.join(os.path.dirname(file_path), "temp.pdf")
        # 将输出PDF写入临时文件
        with open(pdf_file, "wb") as output:
            output_pdf.write(output)
        # 使用subprocess将PDF文件打印到默认打印机
        subprocess.Popen(["lpr", pdf_file], shell=True)
        # 删除临时PDF文件
        os.remove(pdf_file)
    elif file_type == ".txt":
        # 加载文本文件并拆分为行
        with open(file_path, "r") as f:
            lines = f.readlines()
        # 确定要打印的范围
        if print_range:
            start_line, end_line = print_range.split("-")
            start_line = int(start_line) - 1
            end_line = int(end_line)
            print_area = lines[start_line:end_line]
        else:
            # 如果没有选择范围，则打印整个文档
            print_area = lines
            # 创建一个临时文本文件来存储打印内容
        txt_file = os.path.join(os.path.dirname(file_path), "temp.txt")
            # 将打印区域写入临时文本文件
        with open(txt_file, "w") as output:
            for line in print_area:
                output.write(line)
            # 使用subprocess将文本文件打印到默认打印机
        subprocess.Popen(["lpr", txt_file], shell=True)
# 删除临时文本文件
os.remove(txt_file)

# 添加一个新的按钮以导入文档
root = Tk()
root.title("多格式打印")
import_button = Button(root, text="导入文档", command=import_document)
import_button.pack()

def import_data(file_path):
    # 从Excel或CSV文件中导入数据
    if file_path.endswith(".xlsx"):
        data = pd.read_excel(file_path)
    elif file_path.endswith(".csv"):
        data = pd.read_csv(file_path)
    else:
        print("Unsupported file format")
        return None

    return data


def print_cards(data, font, size, leading, tracking, paper_size):
    # 根据纸张大小设置页面大小
    if paper_size == "Letter":
        page_size = letter
    elif paper_size == "A4":
        page_size = A4
    else:
        print("Unsupported paper size")
        return

    # 设置文本样式
    styles = getSampleStyleSheet()
    style_normal = styles["Normal"]
    style_normal.fontName = font
    style_normal.fontSize = size
    style_normal.leading = leading
    style_normal.firstLineIndent = inch / 4
    style_normal.spaceAfter = 0.2 * inch
    style_normal.spaceBefore = 0.2 * inch

    # 创建PDF文件
    filename = f"cards_{paper_size}.pdf"
    doc = SimpleDocTemplate(filename, pagesize=page_size)

    # 初始化元素列表
    elements = []

    # 遍历数据，创建每个牌位的内容
    for i, row in data.iterrows():
        content = []
        for col in row:
            content.append(Paragraph(str(col), style_normal))
        elements.append(Table([content], colWidths=[6 * inch]))

    # 设置表格样式
    t = Table(elements)
    t.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.black),
        ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.black),
        ("BOX", (0, 0), (-1, -1), 0.25, colors.black),
    ]))

    # 添加表格到PDF文件中
    elements.append(t)

    # 构建PDF文件
    doc.build(elements)

# 添加一个设置页面
def settings_page():
    settings_window = Toplevel(root)
    settings_window.title("设置")
    settings_window.geometry("400x300")

    # 添加其他部件
    pass

# 添加选择横向或竖向的OptionMenu
orientation_label = Label(root, text="选择纸张方向")
orientation_label.pack()

orientation_var = StringVar(root)
orientation_var.set("竖向")

orientation_optionmenu = OptionMenu(root, orientation_var, "竖向", "横向")
orientation_optionmenu.pack()

# 创建一个Label和OptionMenu来选择纸张大小
size_label = Label(root, text="选择纸张大小")
size_label.pack()
paper_sizes = ["A5", "A4", "A3", "B5", "B4", "B3", "Letter", "Legal", "Tabloid"]
selected_size = StringVar()
selected_size.set(paper_sizes[0])
size_optionmenu = OptionMenu(root, selected_size, *paper_sizes)
size_optionmenu.pack()

# 创建一个Entry来允许用户输入自定义纸张大小
custom_label = Label(root, text="自定义纸张大小")
custom_label.pack()
custom_entry = Entry(root)
custom_entry.pack()

# 创建一个Label和Entry来选择字体和字号
font_label = Label(root, text="选择字体")
font_label.pack()
font_entry = Entry(root)
font_entry.pack()

size_label = Label(root, text="选择字号")
size_label.pack()
size_entry = Entry(root)
size_entry.pack()

# 创建一个Entry来允许用户输入行距和字距
leading_label = Label(root, text="行距")
leading_label.pack()
leading_entry = Entry(root)
leading_entry.pack()

tracking_label = Label(root, text="字距")
tracking_label.pack()
tracking_entry = Entry(root)
tracking_entry.pack()

# 添加选择横向或竖向的Button
def set_orientation():
    if orientation_var.get() == "竖向":
        size_var.set(size_options[size_var.get()][1])
    else:
        size_var.set(size_options[size_var.get()][0])

orientation_button = Button(root, text="确定", command=set_orientation)
orientation_button.pack()

class Printer:
    def __init__(self, name, address):
        self.name = name
        self.address = address

    def __repr__(self):
        return f"{self.name} ({self.address})"

def print_text(printer_name, text, header=None, footer=None, watermark=None):
    # 连接到CUPS服务器
    conn = cups.Connection()

    # 获取打印机信息
    printers = conn.getPrinters()
    printer_uri = printers[printer_name]['device-uri']

    # 设置打印选项
    job_options = {
        'media': 'A4',
        'fit-to-page': True,
    }

    # 添加页眉
    if header:
        text = header + '\n' + text

    # 添加页脚
    if footer:
        text += '\n' + footer

    # 添加水印
    if watermark:
        text = f'{watermark}\n\n{text}'

    # 添加页码
    text = add_page_numbers(text)

    # 打印文本
    conn.printFile(printer_name, "-", job_options, text.encode('utf-8'))

def add_page_numbers(text):
    # 按行分割文本
    lines = text.strip().split('\n')

    # 计算文本需要占用的页数
    num_pages = len(lines) // 60 + 1

    # 为每一行添加页码
    for i in range(len(lines)):
        line_num = i + 1
        page_num = (line_num - 1) // 60 + 1
        lines[i] = f'{page_num:>3d} - {lines[i]}'

    # 为每一页添加页面底部的页码
    for i in range(num_pages):
        page_num = i + 1
        page_start = i * 60
        page_end = page_start + 60
        lines[page_end - 1] += f'  ---  {page_num}/{num_pages}'

    # 将修改后的文本行连接成一个字符串
    return '\n'.join(lines)

# 例子用法
text_to_print = "这是一个测试打印"
printer_name = "我的打印机"
header = "页眉"
footer = "页脚"
watermark = "保密"
print_text(printer_name, text_to_print, header=header, footer=footer, watermark=watermark)

# 通过CUPS连接打印机
conn = cups.Connection()

# 获取所有打印机
printers = conn.getPrinters()

# 遍历打印机，找到指定名称的打印机
target_printer = None
for printer_name in printers:
    if printer_name.lower() == target_printer_name.lower():
        target_printer = Printer(printer_name, printers[printer_name]['device-uri'])
        break

# 如果没有找到指定名称的打印机，抛出异常
if not target_printer:
    raise ValueError(f"无法找到名称为{target_printer_name}的打印机。")

# 根据打印机类型连接打印机并打印文本
if target_printer.address:
    # 如果是蓝牙打印机，通过蓝牙连接打印机
    subprocess.call(["sudo", "hciconfig", "hci0", "piscan"])
    socket = bluetooth.BluetoothSocket(bluetooth.RFCOMM)
    socket.connect((target_printer.address, 1))
    socket.send(text.encode('utf-8'))
    socket.close()
else:
    # 如果是网络打印机，通过CUPS连接打印机
    job_options = {
        'media': 'A4',
        'fit-to-page': True,
    }
    conn.printFile(target_printer.name, "-", job_options, text.encode('utf-8'))

    # 添加页眉
    if header:
        text = header + '\n' + text

    # 添加页脚
    if footer:
        text += '\n' + footer

    # 添加水印
    if watermark:
        text = f'{watermark}\n\n{text}'

    # 添加页码
    text = add_page_numbers(text)

    # 连接到打印机并打印文本
    if printer_address:
        # 连接蓝牙打印机
        subprocess.call(["sudo", "hciconfig", "hci0", "piscan"])
        socket = bluetooth.BluetoothSocket(bluetooth.RFCOMM)
        socket.connect((printer_address, 1))
        socket.send(text.encode('utf-8'))
        socket.close()
    else:
        # 连接网络打印机
        conn = cups.Connection()
        job_options = {
            'media': 'A4',
            'fit-to-page': True,
        }
        conn.printFile(printer_name, "-", job_options, text.encode('utf-8'))


# 添加支持远程打印功能
    class RemotePrinterSettings(QWidget):
        def __init__(self, remote_printer):
            super().__init__()

            self.remote_printer = remote_printer

            self.host_lineedit = QLineEdit()
            self.host_lineedit.setText(self.remote_printer.host)
            self.port_lineedit = QLineEdit()
            self.port_lineedit.setText(str(self.remote_printer.port))

            self.save_button = QPushButton("Save")
            self.save_button.clicked.connect(self.save_settings)

            layout = QVBoxLayout()
            layout.addWidget(QLabel("Host:"))
            layout.addWidget(self.host_lineedit)
            layout.addWidget(QLabel("Port:"))
            layout.addWidget(self.port_lineedit)
            layout.addWidget(self.save_button)

            self.setLayout(layout)

        def save_settings(self):
            host = self.host_lineedit.text()
            port = int(self.port_lineedit.text())
            self.remote_printer.set_host_port(host, port)
            self.close()

class PrintRemote(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        self.print_btn = QPushButton('打印')
        self.print_btn.clicked.connect(self.print)

        self.connect_btn = QPushButton('连接打印机')
        self.connect_btn.clicked.connect(self.connect)

        self.status_label = QLabel('未连接')

        grid = QGridLayout()
        grid.addWidget(self.print_btn, 0, 0)
        grid.addWidget(self.connect_btn, 0, 1)
        grid.addWidget(self.status_label, 1, 0, 1, 2)

        self.setLayout(grid)

    def print(self):
        pass

    def connect(self):
        pass


class Printer:
    def __init__(self, address):
        self.address = address
        self.socket = None

    def connect(self):
        try:
            self.socket = bluetooth.BluetoothSocket(bluetooth.RFCOMM)
            self.socket.connect((self.address, 1))
            return True
        except:
            return False

    def disconnect(self):
        self.socket.close()

    def print(self, data):
        self.socket.send(data)


class PrinterController:
    def __init__(self):
        self.printers = []

    def add_printer(self, address):
        printer = Printer(address)
        if printer.connect():
            self.printers.append(printer)

    def remove_printer(self, address):
        for printer in self.printers:
            if printer.address == address:
                printer.disconnect()
                self.printers.remove(printer)

    def print(self, address, data):
        for printer in self.printers:
            if printer.address == address:
                printer.print(data)

class RemotePrinter:
        def __init__(self, host, port):
            self.host = host
            self.port = port
            self.socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            self.connected = False

        def connect(self):
            try:
                self.socket.connect((self.host, self.port))
                self.connected = True
            except Exception as e:
                print(f"Error connecting to remote printer: {e}")
                self.connected = False

        def disconnect(self):
            try:
                self.socket.close()
                self.connected = False
            except:
                pass

        def printer(self, data):
            if not self.connected:
                print("Error: Printer is not connected")
                return

            if any(ord(char) > 127 for char in data):
                # 数据包含非ASCII字符，需要采取适当措施
                print("Error: Data contains non-ASCII characters")
                return

            with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
                s.connect((self.host, self.port))
                s.sendall(data.encode('utf-8'))

            print("Print success!")

            try:
                self.socket.sendall(data.encode())
            except Exception as e:
                print(f"Error sending data to remote printer: {e}")
                self.disconnect()
                return

# 初始化远程打印机
remote_printer = RemotePrinter('192.168.0.100', 9999)

def connect():
    try:
        remote_printer.print('连接成功')
        connect_status.config(text='已连接')
        connect_btn.config(state='disabled')
        print_btn.config(state='normal')
    except:
        connect_status.config(text='连接失败')

def print_data():
    data = '要打印的数据'
    remote_printer.print(data)

connect_btn = tk.Button(root, text='连接', command=connect)
print_btn = tk.Button(root, text='打印', state='disabled', command=print_data)
connect_status = tk.Label(root, text='未连接')

connect_btn.pack()
print_btn.pack()
connect_status.pack()


# 创建一个Button来保存设置
def save_settings():
        size = selected_size.get()
        custom_size = custom_entry.get()
        font = font_entry.get()
        size = size_entry.get()
        leading = leading_entry.get()
        tracking = tracking_entry.get()
        settings = {"size": size, "custom_size": custom_size, "font": font, "size": size, "leading": leading, "tracking": tracking}
        with open("settings.pkl", "wb") as f:
            pickle.dump(settings, f)
        messagebox.showinfo("保存成功", "设置已保存！")

save_button = Button(settings_window, text="保存设置", command=save_settings)
save_button.pack()

# 添加一个预览功能
def preview():
    size = selected_size.get()
    orientation = selected_orientation.get()
    custom_size = custom_entry.get()
    font = font_entry.get()
    size = size_entry.get()
    leading = leading_entry.get()
    tracking = tracking_entry.get()

    # 使用先前保存的设置来指定用户选择的自定义纸张大小
    if custom_size:
        size_str = custom_size
    else:
        size_str = size

    # 添加本地字体
    pdfmetrics.registerFont(TTFont('微软雅黑', 'msyh.ttf'))
    pdfmetrics.registerFont(TTFont('黑体', 'simhei.ttf'))

    # 读取系统字体
    fonts_dir = "C:\\Windows\\Fonts\\"
    for font_file in os.listdir(fonts_dir):
        if font_file.endswith(".ttf"):
            font_name = os.path.splitext(font_file)[0]
            try:
                pdfmetrics.registerFont(TTFont(font_name, fonts_dir + font_file))
            except:
                pass

    # 创建PDF文件并添加文本
    filename = "preview.pdf"
    c = canvas.Canvas(filename, pagesize=pagesize)

    # 计算页面可用宽度
    usable_width = width - margin * 2

    # 设置字体、字号、行距、字距
    c.setFont(font, size)
    c.setLeading(leading)
    c.setCharSpace(tracking)

    # 计算文本框可用高度
    usable_height = height - margin * 2 - image_height

    # 将文本框内容分割为多行
    lines = textwrap.wrap(text, width=int(usable_width / (size / 2)))

    # 计算每行的高度
    line_height = c._leading + size

    # 计算文本框实际高度
    text_height = len(lines) * line_height

    # 如果文本框实际高度大于可用高度，自动调整字号和行距
    while text_height > usable_height:
        # 缩小字号
        size -= 1

        # 重新设置字体、字号、行距、字距
        c.setFont(font, size)
        c.setLeading(leading)
        c.setCharSpace(tracking)

        # 计算每行的高度
        line_height = c._leading + size

        # 重新计算文本框实际高度
        lines = textwrap.wrap(text, width=int(usable_width / (size / 2)))
        text_height = len(lines) * line_height

        # 在PDF页面上绘制文本框
    y = height - margin - image_height - text_height
    for line in lines:
        c.drawString(margin, y, line)
        y -= line_height

    # 添加图片
    if image_path:
        img = Image.open(image_path)
        img = img.resize((int(img.width / 2), int(img.height / 2)))
        w, h = img.size
        c.drawImage(image_path, c._xinch(1), c._yinch(1), width=w, height=h)

    # 保存PDF文件
    c.showPage()
    c.save()

    # 打开预览文件
    if os.name == "nt":
        os.startfile(filename)
    elif os.name == "posix":
        subprocess.Popen(["xdg-open", filename])
    else:
        messagebox.showerror("错误", "无法打开预览文件！")


# 添加一个自动调整功能
def auto_adjust():
    size = size_entry.get()
    leading = leading_entry.get()
    tracking = tracking_entry.get()
    text = text_entry.get("1.0", END)

    # 自动调整字号和行距
    size, leading = adjust_text(text, int(size), int(leading), float(tracking))

    # 更新字号和行距
    size_entry.delete(0, END)
    size_entry.insert(0, size)
    leading_entry.delete(0, END)
    leading_entry.insert(0, leading)

# 自动裁剪图片
def auto_crop_image(image_path):
    img = Image.open(image_path)

    # 将图片转换为灰度图像
    img = ImageOps.grayscale(img)

    # 将图片转换为二值图像
    threshold = 128
    img = ImageOps.invert(img)
    img = ImageOps.threshold(img, threshold)

    # 查找图片边框
    border = img.getbbox()

    # 裁剪图片
    if border:
        img = img.crop(border)

    return img

# 在主窗口上添加设置、预览和自动调整按钮
settings_button = Button(root, text="设置", command=settings_page)
settings_button.pack(side=LEFT)

preview_button = Button(root, text="预览", command=preview)
preview_button.pack(side=LEFT)

auto_adjust_button = Button(root, text="自动调整", command=auto_adjust)
auto_adjust_button.pack(side=LEFT)

class InsertImageDialog(QDialog):
    def __init__(self):
        super().__init__()
        self.image_path = None
        self.position_x = 0
        self.position_y = 0
        self.width = 100
        self.height = 100

        self.init_ui()

    def init_ui(self):
        self.setWindowTitle("Insert Image")

        # Create widgets
        self.image_label = QLabel()
        self.image_label.setFixedSize(200, 200)
        self.choose_image_button = QPushButton("Choose Image...")
        self.choose_image_button.clicked.connect(self.choose_image)
        self.position_x_spinbox = QSpinBox()
        self.position_x_spinbox.setRange(0, 10000)
        self.position_x_spinbox.valueChanged.connect(self.update_image_position)
        self.position_y_spinbox = QSpinBox()
        self.position_y_spinbox.setRange(0, 10000)
        self.position_y_spinbox.valueChanged.connect(self.update_image_position)
        self.width_spinbox = QSpinBox()
        self.width_spinbox.setRange(1, 10000)
        self.width_spinbox.valueChanged.connect(self.update_image_size)
        self.height_spinbox = QSpinBox()
        self.height_spinbox.setRange(1, 10000)
        self.height_spinbox.valueChanged.connect(self.update_image_size)
        self.insert_button = QPushButton("Insert")
        self.insert_button.clicked.connect(self.accept)
        self.cancel_button = QPushButton("Cancel")
        self.cancel_button.clicked.connect(self.reject)

        # Layout widgets
        image_layout = QHBoxLayout()
        image_layout.addWidget(self.image_label)
        image_layout.addWidget(self.choose_image_button)

        position_layout = QHBoxLayout()
        position_layout.addWidget(QLabel("Position X:"))
        position_layout.addWidget(self.position_x_spinbox)
        position_layout.addWidget(QLabel("Position Y:"))
        position_layout.addWidget(self.position_y_spinbox)

        size_layout = QHBoxLayout()
        size_layout.addWidget(QLabel("Width:"))
        size_layout.addWidget(self.width_spinbox)
        size_layout.addWidget(QLabel("Height:"))
        size_layout.addWidget(self.height_spinbox)

        button_layout = QHBoxLayout()
        button_layout.addWidget(self.insert_button)
        button_layout.addWidget(self.cancel_button)

        main_layout = QVBoxLayout()
        main_layout.addLayout(image_layout)
        main_layout.addLayout(position_layout)
        main_layout.addLayout(size_layout)
        main_layout.addLayout(button_layout)

        self.setLayout(main_layout)

    def choose_image(self):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        file_name, _ = QFileDialog.getOpenFileName(self, "Choose Image", "",
                                                   "Images (*.png *.xpm *.jpg *.bmp *.gif)", options=options)
        if file_name:
            self.image_path = file_name
            pixmap = QPixmap(self.image_path).scaled(self.image_label.size(), aspectRatioMode=True)
            self.image_label.setPixmap(pixmap)

    def update_image_position(self):
        self.position_x = self.position_x_spinbox.value()
        self.position_y = self.position_y_spinbox.value()

    def update_image_size(self):
        self.width = self.width_spinbox.value()
        self.height = self.height_spinbox.value()

class ImageWidget(QtWidgets.QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.image = None
        self.aspectRatioMode = QtCore.Qt.KeepAspectRatio
        self.sizePolicy = QtWidgets.QSizePolicy(
            QtWidgets.QSizePolicy.Ignored, QtWidgets.QSizePolicy.Ignored)
        self.setMinimumSize(10, 10)

    def setImage(self, image):
        self.image = image
        self.setMinimumSize(1, 1)
        self.update()

    def setAspectRatioMode(self, aspectRatioMode):
        self.aspectRatioMode = aspectRatioMode
        self.update()

    def setSizePolicy(self, sizePolicy):
        self.sizePolicy = sizePolicy
        self.update()

    def sizeHint(self):
        if self.image:
            return self.image.size()
        return QtCore.QSize(10, 10)

    def paintEvent(self, event):
        painter = QtGui.QPainter(self)
        painter.setRenderHint(QtGui.QPainter.Antialiasing, True)
        painter.setRenderHint(QtGui.QPainter.SmoothPixmapTransform, True)
        if self.image:
            target = QtCore.QRectF(self.rect())
            source = QtCore.QRectF(self.image.rect())
            painter.drawImage(target, self.image, source)

# 创建一个ImageWidget对象，并设置图片和大小
image_widget = ImageWidget()
image = QtGui.QPixmap('path/to/image.png')
image_widget.setImage(image)
image_widget.setFixedSize(100, 100)

# 将ImageWidget对象添加到打印牌位的布局中
layout.addWidget(image_widget, row, col, rowspan, colspan)


# 创建一个Button来打印牌位
def print_card():
    size = selected_size.get()
    custom_size = custom_entry.get()
    font = font_entry.get()
    size = size_entry.get()
    leading = leading_entry.get()
    tracking = tracking_entry.get()
    # 使用先前保存的设置来指定用户选择的自定义纸张大小
    if custom_size:
        size_str = custom_size
    else:
        size_str = size
    # 创建PDF文件
    filename = f"card_{size_str}.pdf"

    # 添加本地字体
    pdfmetrics.registerFont(TTFont('微软雅黑', 'msyh.ttf'))
    pdfmetrics.registerFont(TTFont('黑体', 'simhei.ttf'))

    # 读取系统字体
    fonts_dir = "C:\\Windows\\Fonts\\"
    for font_file in os.listdir(fonts_dir):
        if font_file.endswith(".ttf"):
            font_name = os.path.splitext(font_file)[0]
            try:
                pdfmetrics.registerFont(TTFont(font_name, fonts_dir + font_file))
            except:
                pass

        # 创建PDF文件
    c = canvas.Canvas(filename)

    # 设置字体、字号、行距、字距
    c.setFont(font, size)
    c.setLeading(leading)
    c.setCharSpace(tracking)

    # 自动裁剪图片
    def auto_crop_image(image_path):
        img = Image.open(image_path)

        # 将图片转换为灰度图像
        img = ImageOps.grayscale(img)

        # 将图片转换为二值图像
        threshold = 128
        img = ImageOps.invert(img)
        img = ImageOps.threshold(img, threshold)

        # 查找图片边框
        border = img.getbbox()

        # 裁剪图片
        if border:
            img = img.crop(border)

        return img
    # 写入牌位信息
    # ...


    c.save()

    # 打印PDF文件
    if os.name == "nt":
        subprocess.Popen(
            ["C:/Program Files (x86)/Adobe/Acrobat Reader DC/Reader/AcroRd32.exe", "/t", filename, "\\x07"])


# 定义打印机
class Printer:
    def __init__(self, name, address):
        self.name = name
        self.address = address

    def __repr__(self):
        return f"{self.name} ({self.address})"

# 查找蓝牙设备
def discover_devices():
    devices = bluetooth.discover_devices()
    printers = []
    for addr in devices:
        name = bluetooth.lookup_name(addr)
        if "printer" in name.lower():
            printers.append(Printer(name, addr))
    return printers

# 连接打印机
def connect_printer(printer):
    # 连接蓝牙打印机
    subprocess.call(["sudo", "hciconfig", "hci0", "piscan"])
    socket = bluetooth.BluetoothSocket(bluetooth.RFCOMM)
    socket.connect((printer.address, 1))
    # 打印测试页
    socket.send(b"\x1b\x40")  # 复位打印机
    socket.send(b"\x1b\x52\x08")  # 设置字符编码为GB2312
    socket.send("Hello, World!".encode("gb2312"))
    socket.send(b"\n\n")
    socket.send(b"\x1d\x56\x42\x05\x00")  # 切纸
    # 断开蓝牙连接
    socket.close()

# 查找网络打印机
def discover_network_printers():
    conn = cups.Connection()
    printers = []
    for printer in conn.getPrinters():
        printers.append(Printer(printer, conn.getPrinters()[printer]['device-uri']))
    return printers

# 连接网络打印机
def connect_network_printer(printer):
    # 打印测试页
    conn = cups.Connection()
    conn.printFile(printer.name, "test.txt", "", {})

# 创建一个Button来选择图片
def choose_image():
    global image_path
    image_path = askopenfilename()
    image_label.config(text=f"已选择图片：{image_path}")


# 创建一个Label和Button来选择图片
image_label = Label(root, text="请选择图片：")
image_label.pack()
image_button = Button(root, text="选择图片", command=choose_image)
image_button.pack()


# 创建一个Button来添加图片
def add_image():
    size = selected_size.get()
    custom_size = custom_entry.get()
    # 使用先前保存的设置来指定用户选择的自定义纸张大小
    if custom_size:
        size_str = custom_size
    else:
        size_str = size
    # 创建PDF文件
    filename = f"card_{size_str}.pdf"
    c = canvas.Canvas(filename)

    # 添加图片
    if image_path:
        c.drawImage(image_path, inch, inch, width=6 * inch, height=4 * inch)

    c.save()


add_image_button = Button(root, text="添加图片", command=add_image)
add_image_button.pack()


# 创建一个Button来批量打印牌位
def batch_print():
    size = selected_size.get()
    custom_size = custom_entry.get()
    font = font_entry.get()
    size = size_entry.get()
    leading = leading_entry.get()
    tracking = tracking_entry.get()
    # 使用先前保存的设置来指定用户选择的自定义纸张大小
    if custom_size:
        size_str = custom_size
    else:
        size_str = size
    # 创建PDF文件
    filename = f"cards_{size_str}.pdf"

    # 添加本地字体
    pdfmetrics.registerFont(TTFont('微软雅黑', 'msyh.ttf'))
    pdfmetrics.registerFont(TTFont('黑体', 'simhei.ttf'))

    # 读取系统字体
    fonts_dir = "C:\\Windows\\Fonts\\"
    for font_file in os.listdir(fonts_dir):
        if font_file.endswith(".ttf"):
            font_name = os.path.splitext(font_file)[0]
            try:
                pdfmetrics.registerFont(TTFont(font_name, fonts_dir + font_file))
            except:
                pass

    # 创建PDF文件
    c = canvas.Canvas(filename, pagesize=getattr(landscape(selected_size), selected_orientation))

    # 设置字体、字号、行距、字距
    c.setFont(font, size)
    c.setLeading(leading)
    c.setCharSpace(tracking)

    # 写入牌位信息
    # 循环添加牌位
    for i in range(batch_size):
        c.showPage()
        # 写入牌位信息
        # ...

    c.save()

    # 打印PDF文件
    if os.name == "nt":
        subprocess.Popen(["C:/Program Files (x86)/Adobe/Acrobat Reader DC/Reader/AcroRd32.exe", "/t", filename, "\\x07"])

# 创建一个Label和OptionMenu来选择纸张大小
size_label = Label(root, text="选择纸张大小")
size_label.pack()
size_options = ["A5", "A4", "A3", "B5", "B4", "B3", "Letter", "Legal", "Tabloid"]
selected_size = StringVar()
selected_size.set(size_options[0])
size_menu = OptionMenu(root, selected_size, *size_options)
size_menu.pack()

# 创建一个Label和OptionMenu来选择纸张方向
orientation_label = Label(root, text="选择纸张方向")
orientation_label.pack()
orientation_options = ["portrait", "landscape"]
selected_orientation = StringVar()
selected_orientation.set(orientation_options[0])
orientation_menu = OptionMenu(root, selected_orientation, *orientation_options)
orientation_menu.pack()

# 创建一个Label和Scale来选择牌位批量打印数量
batch_label = Label(root, text="批量打印数量")
batch_label.pack()
batch_scale = Scale(root, from_=1, to=10, orient=HORIZONTAL)
batch_scale.pack()

# 创建一个Button来保存设置
def save_settings():
    size = selected_size.get()
    orientation = selected_orientation.get()
    custom_size = custom_entry.get()
    font = font_entry.get()
    size = size_entry.get()
    leading = leading_entry.get()
    tracking = tracking_entry.get()
    batch_size = batch_scale.get()
    settings = {"size": size, "orientation": orientation, "custom_size": custom_size, "font": font, "size": size, "leading": leading, "tracking": tracking, "batch_size": batch_size}
    with open("settings.pkl", "wb") as f:
        pickle.dump(settings, f)

save_button = Button(root, text="保存设置", command=save_settings)
save_button.pack()

# 创建一个Button来打印牌位
def print_card():
    size = selected_size.get()
    orientation = selected_orientation.get()
    custom_size = custom_entry.get()
    font = font_entry.get()
    size = size_entry.get()
    leading = leading_entry.get()
    tracking = tracking_entry.get()
    # 使用先前保存的设置来指定用户选择的自定义纸张大小
    if custom_size:
        size_str = custom_size
    else:
        size_str = size
    # 创建PDF文件
    filename = f"card_{size_str}.pdf"

    # 添加本地字体
    pdfmetrics.registerFont(TTFont('微软雅黑', 'msyh.ttf'))
    pdfmetrics.registerFont(TTFont('黑体', 'simhei.ttf'))

    # 读取系统字体
    fonts_dir = "C:\\Windows\\Fonts\\"
    for font_file in os.listdir(fonts_dir):
        if font_file.endswith(".ttf"):
            font_name = os.path.splitext(font_file)[0]
            try:
                pdfmetrics.registerFont(TTFont(font_name, fonts_dir + font_file))
            except:
                pass

        # 创建PDF文件
    c = canvas.Canvas(filename, pagesize=getattr(landscape(selected_size), selected_orientation))

    # 设置字体、字号、行距、字距
    c.setFont(font, size)
    c.setLeading(leading)
    c.setCharSpace(tracking)

    # 循环添加牌位
    for i in range(batch_size):
        if i % len(image_list) == 0:
            c.showPage()
        if i < len(image_list):
            c.drawImage(ImageReader(image_list[i]), 0, 0, width=100, height=100, mask='auto')
    else:
     # 写入牌位信息
     # ...

     c.save()

    # 打印PDF文件
    if os.name == "nt":
        subprocess.Popen(
            ["C:/Program Files (x86)/Adobe/Acrobat Reader DC/Reader/AcroRd32.exe", "/t", filename, "\\x07"])
        # 添加图片
        if image_list:
            for j, image_path in enumerate(image_list):
                if i == j % batch_size:
                    img = Image.open(image_path)
                    img = img.resize((int(img.width / 2), int(img.height / 2)))
                    w, h = img.size
                    c.drawImage(image_path, c._xinch(1), c._yinch(1), width=w, height=h)

        # 保存PDF文件
        c.save()

    # 显示打印成功的提示
    messagebox.showinfo("打印成功", f"{batch_size} 个牌位已打印！")

# 更新Canvas上的图片
update_canvas()

class CardWidget(QWidget):
    def __init__(self, card_data, font_size=12):
        super().__init__()

        self.card_data = card_data
        self.font_size = font_size

        self.labels = []

        # 创建标签
        for row in self.card_data:
            label = QLabel(row, self)
            label.setStyleSheet("font-size: {}pt".format(self.font_size))
            self.labels.append(label)

        # 设置布局
        self.layout = QVBoxLayout(self)
        for label in self.labels:
            self.layout.addWidget(label)

        # 根据文本内容的长度和字体大小调整布局
        self.adjust_size()

    def adjust_size(self):
        # 获取最长的标签文本长度
        max_length = max([len(label.text()) for label in self.labels])

        # 计算标签的最小宽度
        font_metrics = QFontMetrics(self.font())
        min_width = font_metrics.width("W" * max_length)

        # 设置标签的最小宽度
        for label in self.labels:
            label.setMinimumWidth(min_width)

    def set_font_size(self, font_size):
        self.font_size = font_size

        # 更新标签样式
        for label in self.labels:
            label.setStyleSheet("font-size: {}pt".format(self.font_size))

        # 根据文本内容的长度和字体大小调整布局
        self.adjust_size()

# 定义自动纠错和矫正函数
def auto_correct(text):
    # 定义需要矫正的常见错误和对应的正常词语
    correction_dict = {"佛光照照": "佛光注照", "阳上拜荐": "阳上拜见", "长生鹿位": "长生禄位"}
    # 将所有需要矫正的错误匹配出来，并替换为正确词语
    for error, correction in correction_dict.items():
        pattern = re.compile(error)
        text = re.sub(pattern, correction, text)
    return text

def print_stats(data):
    # 打印统计和报表
    num_cards = len(data)
    num_pages = int(num_cards / 8) + (1 if num_cards % 8 != 0 else 0)
    num_words = data.applymap(lambda x: len(str(x)).split()).sum().sum()
    print(f"Number of cards: {num_cards}")
    print(f"Number of pages: {num_pages}")
    print(f"Number of words: {num_words}")

    # 导出为CSV文件
    data.to_csv("cards.csv", index=False)

if __name__ == '__main__':
    app = QApplication([])
    card_data = ['佛光注照（）长生禄位', '佛力超荐（）往生莲位', '阳上:（）拜荐']
    widget = CardWidget(card_data)
    widget.show()
    app.exec_()

# 创建一个Button来退出程序
def quit_program():
      root.destroy
